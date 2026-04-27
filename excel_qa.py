from __future__ import annotations

import json
import logging
import os
import re
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import es_search
from pdf_qa import DEFAULT_BASE_URL, DEFAULT_MODEL, build_openai_client, encode_texts, extract_response_text
from storage import (
    get_excel_filter_enums,
    get_excel_policy_chunks_by_positions,
    get_document,
    list_excel_chunks_for_embedding,
    list_excel_chunks_for_search_index,
    list_ready_excel_documents,
    save_excel_policy_index,
    save_excel_chunk_vector,
    search_excel_policy_chunks,
    sqlite_vec_available,
    update_document_ingestion_config,
    vector_search_excel_policy_chunks,
)

logger = logging.getLogger(__name__)


DEFAULT_EXCEL_CHUNK_SIZE = 1000
DEFAULT_EXCEL_CHUNK_OVERLAP = 100
MAX_CONTEXT_CHUNKS = 10
MAX_CONTEXT_POLICIES = 6
EXCEL_RETRIEVAL_SOURCE_TOP_K = 5
GENERAL_CHAT_PATTERNS = (
    "你好",
    "您好",
    "hello",
    "hi",
    "谢谢",
    "感谢",
)
EXCEL_QUERY_CLASSIFIER_MODEL = os.environ.get("EXCEL_QUERY_CLASSIFIER_MODEL", "").strip()
EXCEL_VECTOR_INDEX_ENABLED = os.environ.get("EXCEL_VECTOR_INDEX_ENABLED", "1").strip() != "0"


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_xls_source(path_or_bytes: str | Path | bytes) -> bool:
    if isinstance(path_or_bytes, bytes):
        return path_or_bytes.startswith(b"\xd0\xcf\x11\xe0")
    return Path(path_or_bytes).suffix.lower() == ".xls"


def _read_excel_rows(path_or_bytes: str | Path | bytes) -> tuple[str, list[list[str]]]:
    if _is_xls_source(path_or_bytes):
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError("解析 .xls 文件需要安装 xlrd，请先执行 pip install -r requirements.txt") from exc

        if isinstance(path_or_bytes, bytes):
            workbook = xlrd.open_workbook(file_contents=path_or_bytes)
        else:
            workbook = xlrd.open_workbook(str(path_or_bytes))
        sheet = workbook.sheet_by_index(0)
        rows = [
            [_cell_to_text(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]
            for row_index in range(sheet.nrows)
        ]
        return sheet.name, rows

    source = BytesIO(path_or_bytes) if isinstance(path_or_bytes, bytes) else path_or_bytes
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = [
            [_cell_to_text(value) for value in values]
            for values in sheet.iter_rows(values_only=True)
        ]
        return sheet.title, rows
    finally:
        workbook.close()


def _detect_header_row_index(rows: list[list[str]]) -> int:
    candidates = rows[:20]
    best_index = 0
    best_score = -1
    for index, row in enumerate(candidates):
        values = [value.strip() for value in row if value.strip()]
        if not values:
            continue
        non_empty_count = len(values)
        next_non_empty_count = (
            len([value.strip() for value in rows[index + 1] if value.strip()])
            if index + 1 < len(rows)
            else 0
        )
        known_header_hits = sum(
            1
            for value in values
            if value in {"序号", "标题", "级别", "政策发文号", "政策发文字号", "政策原文", "正文", "内容"}
        )
        score = non_empty_count * 10 + known_header_hits * 8 + min(next_non_empty_count, non_empty_count)
        if non_empty_count == 1 and next_non_empty_count > 1:
            score -= 20
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def parse_excel_preview(path_or_bytes: str | Path | bytes, sample_limit: int = 3) -> dict[str, Any]:
    sheet_name, rows = _read_excel_rows(path_or_bytes)
    if not rows:
        raise ValueError("Excel 首行必须是表头")

    header_row_index = _detect_header_row_index(rows)
    columns = rows[header_row_index]
    if not any(columns):
        raise ValueError("Excel 表头为空")

    sample_rows: list[dict[str, str]] = []
    row_count = 0
    for values in rows[header_row_index + 1 :]:
        if not values or not any(value.strip() for value in values):
            continue
        row_count += 1
        if len(sample_rows) < sample_limit:
            sample_rows.append(
                {
                    column: values[index] if index < len(values) else ""
                    for index, column in enumerate(columns)
                    if column
                }
            )

    return {
        "columns": [column for column in columns if column],
        "sample_rows": sample_rows,
        "row_count": row_count,
        "sheet_name": sheet_name,
        "header_row": header_row_index + 1,
    }


def guess_excel_config(columns: list[str]) -> dict[str, Any]:
    def first_match(candidates: list[str], fallback: str = "") -> str:
        lowered = [(column, column.lower()) for column in columns]
        for keyword in candidates:
            keyword_lower = keyword.lower()
            for column, column_lower in lowered:
                if keyword_lower in column_lower:
                    return column
        return fallback

    title_field = first_match(["标题", "名称", "政策名称", "title"], columns[0] if columns else "")
    content_field = first_match(["政策原文", "正文", "内容", "content", "text"], columns[-1] if columns else "")
    filter_field = first_match(["级别", "地区", "类型", "分类", "level"], "")
    source_field = first_match(["政策发文字号", "发文字号", "文号", "来源", "source"], "")
    ignore_fields = [column for column in columns if column in {"序号", "编号", "id", "ID"}]
    display_fields = [
        field
        for field in [title_field, filter_field, source_field]
        if field
    ]

    return {
        "document_mode": "row_as_document",
        "title_field": title_field,
        "content_fields": [content_field] if content_field else [],
        "filter_fields": [filter_field] if filter_field else [],
        "source_fields": [source_field] if source_field else [],
        "display_fields": display_fields,
        "ignore_fields": ignore_fields,
        "chunking": {
            "enabled": True,
            "strategy": "fixed_overlap",
            "fallback_chunk_size": DEFAULT_EXCEL_CHUNK_SIZE,
            "overlap": DEFAULT_EXCEL_CHUNK_OVERLAP,
        },
    }


def normalize_excel_config(config: dict[str, Any], columns: list[str]) -> dict[str, Any]:
    column_set = set(columns)

    def one(field_name: str) -> str:
        value = str(config.get(field_name) or "").strip()
        if value not in column_set:
            raise ValueError(f"{field_name} 不在 Excel 表头中：{value}")
        return value

    def many(field_name: str) -> list[str]:
        raw = config.get(field_name) or []
        if isinstance(raw, str):
            raw = [item.strip() for item in re.split(r"[,，\n]", raw) if item.strip()]
        values: list[str] = []
        for value in raw:
            normalized = str(value or "").strip()
            if not normalized:
                continue
            if normalized not in column_set:
                raise ValueError(f"{field_name} 包含不存在的表头：{normalized}")
            if normalized not in values:
                values.append(normalized)
        return values

    title_field = one("title_field")
    content_fields = many("content_fields")
    if not content_fields:
        raise ValueError("至少需要选择一个正文字段")

    chunking = config.get("chunking") or {}
    chunk_size = int(chunking.get("fallback_chunk_size") or DEFAULT_EXCEL_CHUNK_SIZE)
    overlap = int(chunking.get("overlap") or DEFAULT_EXCEL_CHUNK_OVERLAP)
    chunk_size = max(200, min(5000, chunk_size))
    overlap = max(0, min(chunk_size - 1, overlap))

    return {
        "document_mode": "row_as_document",
        "title_field": title_field,
        "content_fields": content_fields,
        "filter_fields": many("filter_fields"),
        "source_fields": many("source_fields"),
        "display_fields": many("display_fields") or [title_field],
        "ignore_fields": many("ignore_fields"),
        "chunking": {
            "enabled": True,
            "strategy": "fixed_overlap",
            "fallback_chunk_size": chunk_size,
            "overlap": overlap,
        },
    }


def fixed_overlap_chunks(text: str, chunk_size: int, overlap: int) -> list[str]:
    normalized = "\n".join(line.strip() for line in str(text or "").splitlines() if line.strip())
    if not normalized:
        return []
    if len(normalized) <= chunk_size:
        return [normalized]

    chunks: list[str] = []
    start = 0
    step = max(1, chunk_size - overlap)
    while start < len(normalized):
        chunk = normalized[start : start + chunk_size].strip()
        if chunk:
            chunks.append(chunk)
        if start + chunk_size >= len(normalized):
            break
        start += step
    return chunks


def _join_labeled_fields(row: dict[str, str], fields: list[str]) -> str:
    parts = []
    for field in fields:
        value = row.get(field, "").strip()
        if value:
            parts.append(f"{field}：{value}")
    return "\n".join(parts)


def build_excel_chunk_embeddings(document_id: int) -> tuple[int, str]:
    if not EXCEL_VECTOR_INDEX_ENABLED:
        return 0, "Excel 向量索引已通过 EXCEL_VECTOR_INDEX_ENABLED=0 关闭。"
    if not sqlite_vec_available():
        return 0, "sqlite-vec 不可用，Excel 当前仅使用全文索引。"

    chunks = list_excel_chunks_for_embedding(document_id)
    if not chunks:
        return 0, "没有可向量化的 Excel chunk。"

    processed = 0
    total = len(chunks)
    for start in range(0, total, 16):
        batch = chunks[start : start + 16]
        texts = [item["search_text"] for item in batch]
        embeddings = encode_texts(texts)
        for item, embedding in zip(batch, embeddings):
            save_excel_chunk_vector(
                chunk_id=int(item["chunk_id"]),
                policy_id=int(item["policy_id"]),
                document_id=document_id,
                embedding=embedding,
            )
        processed += len(batch)
        progress = 75 + int((processed / max(total, 1)) * 24)
        update_document_ingestion_config(
            document_id,
            config=get_document(document_id).get("config") if get_document(document_id) else {},
            status="processing",
            progress=progress,
            detail=f"正在生成 Excel 向量索引：{processed}/{total} 个片段。",
        )
    return processed, f"Excel 混合索引完成：{processed} 个向量。"


def ingest_excel_file(document_id: int, config: dict[str, Any]) -> dict[str, Any]:
    document = get_document(document_id)
    if not document:
        raise ValueError("文档不存在")
    if document.get("file_type") != "excel":
        raise ValueError("该文档不是 Excel 文件")

    preview = parse_excel_preview(document["storage_path"])
    normalized_config = normalize_excel_config(config, preview["columns"])
    update_document_ingestion_config(
        document_id,
        config=normalized_config,
        row_count=0,
        status="processing",
        progress=10,
        detail="正在解析 Excel 并建立全文索引。",
    )

    policies: list[dict[str, Any]] = []
    sheet_name, raw_rows = _read_excel_rows(document["storage_path"])
    header_row_index = _detect_header_row_index(raw_rows)
    header_values = raw_rows[header_row_index] if raw_rows else []
    columns = [_cell_to_text(value) for value in header_values]
    chunking = normalized_config["chunking"]
    chunk_size = int(chunking["fallback_chunk_size"])
    overlap = int(chunking["overlap"])

    for source_row, values in enumerate(raw_rows[header_row_index + 1 :], start=header_row_index + 2):
        row = {
            column: values[index] if index < len(values) else ""
            for index, column in enumerate(columns)
            if column
        }
        if not any(row.values()):
            continue

        title = row.get(normalized_config["title_field"], "").strip() or f"第 {source_row} 行"
        content = "\n".join(
            row.get(field, "").strip()
            for field in normalized_config["content_fields"]
            if row.get(field, "").strip()
        )
        if not content:
            continue

        metadata = {
            field: row.get(field, "").strip()
            for field in [*normalized_config["filter_fields"], *normalized_config["source_fields"]]
            if row.get(field, "").strip()
        }
        metadata.update(
            {
                "title": title,
                "source_file": document["file_name"],
                "source_sheet": sheet_name,
                "source_row": source_row,
            }
        )
        metadata_text = _join_labeled_fields(
            row,
            [*normalized_config["filter_fields"], *normalized_config["source_fields"]],
        )
        search_text = "\n".join(
            part
            for part in [
                f"标题：{title}",
                metadata_text,
                f"正文：{content}",
            ]
            if part
        )
        chunks = []
        for chunk_index, chunk_text in enumerate(
            fixed_overlap_chunks(content, chunk_size=chunk_size, overlap=overlap),
            start=1,
        ):
            chunk_search_text = "\n".join(
                part
                for part in [
                    f"标题：{title}",
                    metadata_text,
                    f"正文片段：{chunk_text}",
                ]
                if part
            )
            chunks.append(
                {
                    "chunk_index": chunk_index,
                    "chunk_text": chunk_text,
                    "search_text": chunk_search_text,
                    "metadata": metadata,
                }
            )

        policies.append(
            {
                "title": title,
                "content": content,
                "metadata": metadata,
                "search_text": search_text,
                "source_row": source_row,
                "chunks": chunks,
            }
        )

    indexed_count = save_excel_policy_index(
        document_id,
        config=normalized_config,
        policies=policies,
    )
    es_count = 0
    if es_search.enabled():
        try:
            es_search.delete_document("excel", document_id)
            es_count = es_search.index_excel_chunks(list_excel_chunks_for_search_index(document_id))
        except Exception as exc:  # noqa: BLE001
            logger.warning("[excel.es.index.failed] document_id=%s error=%s", document_id, exc)

    vector_count = 0
    vector_detail = ""
    try:
        vector_count, vector_detail = build_excel_chunk_embeddings(document_id)
    except Exception as exc:  # noqa: BLE001
        vector_detail = f"向量索引失败，已保留全文索引：{exc}"

    final_detail = (
        f"Excel 混合索引完成：{indexed_count} 条记录、{vector_count} 个向量"
        f"{f'、{es_count} 个 ES 文档' if es_count else ''}。"
        if vector_count
        else f"Excel 全文索引完成：{indexed_count} 条记录"
        f"{f'、{es_count} 个 ES 文档' if es_count else ''}。{vector_detail}"
    )
    update_document_ingestion_config(
        document_id,
        config=normalized_config,
        row_count=indexed_count,
        status="done",
        progress=100,
        detail=final_detail,
    )
    return {
        "document": get_document(document_id),
        "indexed_count": indexed_count,
        "vector_count": vector_count,
        "config": normalized_config,
    }


def classify_policy_query_by_rule(question: str) -> dict[str, Any]:
    normalized = " ".join(str(question or "").strip().split())
    if not normalized:
        return {
            "intent": "unclear",
            "need_retrieval": False,
            "query_type": "unclear",
            "keywords": [],
            "filters": {},
            "preferred_result": "clarify",
        }

    if normalized.lower() in GENERAL_CHAT_PATTERNS or len(normalized) <= 3:
        return {
            "intent": "general_chat",
            "need_retrieval": False,
            "query_type": "general_chat",
            "keywords": [],
            "filters": {},
            "preferred_result": "normal_reply",
        }

    filters: dict[str, str | None] = {}
    doc_no_match = re.search(r"[\u4e00-\u9fff]{1,12}〔\d{4}〕[^，。；\s]+号", normalized)
    if doc_no_match:
        filters["政策发文字号"] = doc_no_match.group(0)

    level_match = re.search(r"(国家级|省级|市级|区级|县级|苏州市级|江苏省级)", normalized)
    if level_match:
        filters["级别"] = level_match.group(1)

    stop_words = {
        "我",
        "想",
        "知道",
        "我想知道",
        "对于",
        "有没有",
        "什么",
        "好",
        "好的",
        "政策",
        "请问",
        "帮我",
        "查一下",
    }
    keywords = [
        token
        for token in re.split(r"[\s,，。；;？?的]+", normalized)
        if token and token not in stop_words
    ]
    intent = "doc_lookup" if doc_no_match else "policy_search"
    if any(word in normalized for word in ("对比", "比较", "区别")):
        intent = "policy_compare"

    return {
        "intent": intent,
        "need_retrieval": True,
        "query_type": "policy_qa",
        "keywords": keywords[:8] or [normalized],
        "filters": filters,
        "preferred_result": "summary_with_sources",
    }


def _rrf_merge(
    sources: list[tuple[str, list[dict[str, Any]]]],
    k: int = 60,
) -> list[dict[str, Any]]:
    scores: dict[int, float] = {}
    merged: dict[int, dict[str, Any]] = {}
    for source_name, hits in sources:
        for rank, hit in enumerate(hits, start=1):
            cid = int(hit["chunk_id"])
            scores[cid] = scores.get(cid, 0) + 1.0 / (k + rank)
            if cid not in merged:
                merged[cid] = {**hit, "retrieval_sources": []}
            if source_name not in merged[cid]["retrieval_sources"]:
                merged[cid]["retrieval_sources"].append(source_name)
    return [merged[cid] for cid in sorted(scores, key=lambda c: -scores[c])]


def hybrid_search_excel_chunks(
    question: str,
    search_query: str,
    *,
    document_id: int | None,
    filters: dict[str, str],
    top_k: int = MAX_CONTEXT_CHUNKS,
    backend: str | None = None,
) -> list[dict[str, Any]]:
    backend = (backend or "sqlite").strip().lower()
    use_sqlite = backend in {"sqlite", "default", "hybrid", "hybrid_es"}
    use_es = backend in {"es", "elasticsearch", "opensearch", "hybrid", "hybrid_es"}
    es_query = question

    if use_es and not use_sqlite:
        try:
            es_hits = es_search.search_excel_chunks(
                es_query,
                document_id=document_id,
                filters=filters,
                top_k=max(top_k, EXCEL_RETRIEVAL_SOURCE_TOP_K),
            )
            if not es_hits and es_query != question:
                es_hits = es_search.search_excel_chunks(
                    question,
                    document_id=document_id,
                    filters=filters,
                    top_k=max(top_k, EXCEL_RETRIEVAL_SOURCE_TOP_K),
                )
        except Exception as exc:  # noqa: BLE001
            logger.warning("[excel.es.search.failed] question=%r error=%s", question[:80], exc)
            es_hits = []
        return es_hits[:top_k]

    fts_hits = search_excel_policy_chunks(
        search_query,
        document_id=document_id,
        filters=filters,
        top_k=EXCEL_RETRIEVAL_SOURCE_TOP_K,
    )
    if not fts_hits and search_query != question:
        fts_hits = search_excel_policy_chunks(
            question,
            document_id=document_id,
            filters=filters,
            top_k=EXCEL_RETRIEVAL_SOURCE_TOP_K,
        )

    vector_hits: list[dict[str, Any]] = []
    if sqlite_vec_available():
        try:
            embeddings = encode_texts([question])
            if embeddings:
                vector_hits = vector_search_excel_policy_chunks(
                    embeddings[0],
                    document_id=document_id,
                    filters=filters,
                    top_k=EXCEL_RETRIEVAL_SOURCE_TOP_K,
                )
        except Exception as exc:  # noqa: BLE001
            logger.warning("[excel.vector.search.failed] question=%r error=%s", question[:80], exc)
            vector_hits = []

    es_hits: list[dict[str, Any]] = []
    if use_es and es_search.enabled():
        try:
            es_hits = es_search.search_excel_chunks(
                es_query,
                document_id=document_id,
                filters=filters,
                top_k=EXCEL_RETRIEVAL_SOURCE_TOP_K,
            )
            if not es_hits and es_query != question:
                es_hits = es_search.search_excel_chunks(
                    question,
                    document_id=document_id,
                    filters=filters,
                    top_k=EXCEL_RETRIEVAL_SOURCE_TOP_K,
                )
        except Exception as exc:  # noqa: BLE001
            logger.warning("[excel.es.search.failed] question=%r error=%s", question[:80], exc)
            es_hits = []

    source_hits: list[tuple[str, list[dict[str, Any]]]] = [
        ("fts", fts_hits),
        ("vector", vector_hits),
    ]
    if use_es:
        source_hits.append(("elasticsearch", es_hits))

    return _rrf_merge(source_hits)[:top_k]


def _build_classifier_prompt(filter_enums: dict[str, list[str]] | None) -> str:
    lines = [
        "你是政策检索助手，根据用户问题提取检索参数，只输出 JSON。",
        "字段：need_retrieval(bool), keywords(list[str]), filters(dict), intent(str), query_type(str), preferred_result(str)",
        "intent 枚举：policy_search / policy_compare / doc_lookup / general_chat / unclear",
        "",
    ]
    if filter_enums:
        lines.append("当前政策库可用过滤字段（filters 只能从以下值中选，不确定时留 {}）：")
        for field, values in filter_enums.items():
            lines.append(f"  {field}：{values}")
        lines.append("")
    lines += [
        "要求：",
        "- keywords 提取用于全文检索的实质名词，保留字母等级标识（如 E类、A级、B轮）",
        "- filters 严格使用上述枚举值，语义相近但不确定时宁可不过滤",
        "- need_retrieval 为 false 时 keywords 和 filters 留空",
    ]
    return "\n".join(lines)


def _extract_json_object(text: str) -> dict[str, Any] | None:
    stripped = text.strip()
    try:
        result = json.loads(stripped)
        if isinstance(result, dict):
            return result
    except Exception:
        pass
    end = stripped.rfind("}")
    if end == -1:
        return None
    depth = 0
    for i in range(end, -1, -1):
        ch = stripped[i]
        if ch == "}":
            depth += 1
        elif ch == "{":
            depth -= 1
            if depth == 0:
                try:
                    return json.loads(stripped[i : end + 1])
                except Exception:
                    return None
    return None


def classify_policy_query(
    question: str,
    base_url: str = DEFAULT_BASE_URL,
    filter_enums: dict[str, list[str]] | None = None,
) -> dict[str, Any]:
    if not EXCEL_QUERY_CLASSIFIER_MODEL:
        return classify_policy_query_by_rule(question)

    try:
        client = build_openai_client(base_url)
        response = client.chat.completions.create(
            model=EXCEL_QUERY_CLASSIFIER_MODEL,
            messages=[
                {"role": "system", "content": _build_classifier_prompt(filter_enums)},
                {"role": "user", "content": question},
            ],
            max_tokens=512,
            temperature=0,
        )
        text = extract_response_text(response.choices[0].message.content)
        payload = _extract_json_object(text)
        if isinstance(payload, dict) and "need_retrieval" in payload:
            payload.setdefault("keywords", [])
            payload.setdefault("filters", {})
            payload.setdefault("preferred_result", "summary_with_sources")
            payload.setdefault("query_type", "policy_qa")
            payload.setdefault("intent", "policy_search")
            logger.info("[excel.classify.llm] question=%r filters=%s keywords=%s",
                        question[:80], payload["filters"], payload["keywords"])
            return payload
    except Exception as exc:
        logger.warning("[excel.classify.llm.failed] question=%r error=%s", question[:80], exc)

    result = classify_policy_query_by_rule(question)
    logger.info("[excel.classify.rule] question=%r filters=%s keywords=%s",
                question[:80], result["filters"], result["keywords"])
    return result


def group_chunks_by_policy(chunks: list[dict[str, Any]], top_n: int = MAX_CONTEXT_POLICIES) -> list[dict[str, Any]]:
    grouped: dict[int, dict[str, Any]] = {}
    for chunk in chunks:
        policy_id = int(chunk["policy_id"])
        if policy_id not in grouped:
            grouped[policy_id] = {
                "policy_id": policy_id,
                "document_id": int(chunk["document_id"]),
                "document_file_name": chunk["document_file_name"],
                "document_display_name": chunk["document_display_name"],
                "title": chunk["title"],
                "source_row": int(chunk["source_row"]),
                "metadata": chunk.get("metadata") or {},
                "chunks": [],
            }
        grouped[policy_id]["chunks"].append(chunk)
    return list(grouped.values())[:top_n]


def expand_chunks_with_neighbors(
    chunks: list[dict[str, Any]],
    *,
    document_id: int | None,
    radius: int = 1,
) -> list[dict[str, Any]]:
    if radius <= 0 or not chunks:
        return chunks

    original_by_id = {int(chunk["chunk_id"]): chunk for chunk in chunks}
    positions: list[tuple[int, int]] = []
    for chunk in chunks:
        policy_id = int(chunk["policy_id"])
        chunk_index = int(chunk["chunk_index"])
        for offset in range(-radius, radius + 1):
            positions.append((policy_id, chunk_index + offset))

    neighbors = get_excel_policy_chunks_by_positions(positions, document_id=document_id)
    by_position = {
        (int(chunk["policy_id"]), int(chunk["chunk_index"])): chunk
        for chunk in neighbors
    }

    expanded: list[dict[str, Any]] = []
    seen: set[int] = set()
    for chunk in chunks:
        policy_id = int(chunk["policy_id"])
        chunk_index = int(chunk["chunk_index"])
        for offset in range(-radius, radius + 1):
            candidate = by_position.get((policy_id, chunk_index + offset))
            if not candidate:
                continue
            chunk_id = int(candidate["chunk_id"])
            if chunk_id in seen:
                continue
            seen.add(chunk_id)
            if chunk_id in original_by_id:
                original = original_by_id[chunk_id]
                candidate = {
                    **candidate,
                    "retrieval_sources": original.get("retrieval_sources", []),
                    "rank": original.get("rank"),
                }
            else:
                candidate = {**candidate, "retrieval_sources": ["neighbor"]}
            expanded.append(candidate)

    return expanded


def build_excel_context(question: str, policies: list[dict[str, Any]]) -> str:
    blocks = [f"用户问题：\n{question}", "以下是检索到的相关政策片段："]
    for index, policy in enumerate(policies, start=1):
        metadata = policy.get("metadata") or {}
        source_no = metadata.get("政策发文字号") or metadata.get("发文字号") or metadata.get("文号") or ""
        level = metadata.get("级别") or metadata.get("地区") or metadata.get("类型") or ""
        snippets = "\n".join(
            f"- {chunk['chunk_text']}"
            for chunk in policy.get("chunks", [])[:3]
        )
        blocks.append(
            "\n".join(
                part
                for part in [
                    f"【政策{index}】",
                    f"标题：{policy['title']}",
                    f"级别：{level}" if level else "",
                    f"发文字号：{source_no}" if source_no else "",
                    f"来源：{policy['document_file_name']} 第 {policy['source_row']} 行",
                    "命中片段：",
                    snippets,
                ]
                if part
            )
        )

    blocks.append(
        "\n".join(
            [
                "请基于以上政策片段回答用户问题。",
                "要求：",
                "1. 检索片段是候选材料，可能包含不相关内容；请先判断相关性，忽略主题不相关的政策。",
                "2. 只基于相关政策内容回答，不要编造补贴金额、条件、日期。",
                "3. 每一点尽量标注来源政策标题和发文字号；没有发文字号时标注来源文件和行号。",
                "4. 如果资料不足，请说明“当前资料中未检索到明确依据”。",
                "5. 输出结构清晰，优先用分点或表格。",
            ]
        )
    )
    return "\n\n".join(blocks)


def build_excel_answer_request(
    question: str,
    conversation_history: list[dict[str, Any]],
    document_id: int | None = None,
    base_url: str = DEFAULT_BASE_URL,
    max_history_messages: int = 8,
    retrieval_backend: str | None = None,
) -> dict[str, Any] | None:
    filter_enums = get_excel_filter_enums(document_id) if document_id is not None else {}
    query_plan = classify_policy_query(question, base_url=base_url, filter_enums=filter_enums)
    if not query_plan.get("need_retrieval"):
        return {
            "query_plan": query_plan,
            "client": build_openai_client(base_url),
            "messages": [
                {"role": "system", "content": "你是一个简洁的中文助手。"},
                {"role": "user", "content": question},
            ],
            "policies": [],
            "chunks": [],
            "routed_documents": [],
            "answer_sources": [],
        }

    search_query = question
    applied_filters = query_plan.get("filters") or {}
    chunks = hybrid_search_excel_chunks(
        question,
        search_query,
        document_id=document_id,
        filters=applied_filters,
        top_k=MAX_CONTEXT_CHUNKS,
        backend=retrieval_backend,
    )
    logger.info("[excel.retrieve] question=%r filters=%s chunks=%d",
                question[:80], applied_filters, len(chunks))
    if not chunks:
        return None

    chunks = expand_chunks_with_neighbors(chunks, document_id=document_id)
    policies = group_chunks_by_policy(chunks)
    routed_by_id = {int(document["id"]): document for document in list_ready_excel_documents()}
    routed_documents = [
        routed_by_id[document_id]
        for document_id in dict.fromkeys(int(policy["document_id"]) for policy in policies)
        if document_id in routed_by_id
    ]
    context = build_excel_context(question, policies)
    _budget = 12_000
    _used = 0
    history: list[dict[str, Any]] = []
    for msg in reversed(conversation_history[-max_history_messages:]):
        if msg.get("role") not in {"user", "assistant"} or not msg.get("content"):
            continue
        cost = len(msg["content"])
        if _used + cost > _budget:
            break
        history.insert(0, {"role": msg["role"], "content": msg["content"]})
        _used += cost
    messages = [
        {
            "role": "system",
            "content": "你是政策 Excel 检索问答助手，必须基于给定检索片段回答，并保留来源。",
        },
        *history,
        {"role": "user", "content": context},
    ]
    answer_sources = [
        {
            "type": "excel_policy",
            "document_id": policy["document_id"],
            "policy_id": policy["policy_id"],
            "title": policy["title"],
            "source_row": policy["source_row"],
            "metadata": policy.get("metadata") or {},
            "snippets": [chunk["chunk_text"] for chunk in policy.get("chunks", [])[:3]],
        }
        for policy in policies
    ]
    return {
        "query_plan": query_plan,
        "client": build_openai_client(base_url),
        "messages": messages,
        "policies": policies,
        "chunks": chunks,
        "routed_documents": routed_documents,
        "answer_sources": answer_sources,
    }


def build_excel_summary(model: str, usage: dict[str, Any], policies: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "model": model,
        "policy_count_sent": len(policies),
        "sources_sent": [
            {
                "document_id": policy["document_id"],
                "document_name": policy["document_file_name"],
                "policy_id": policy["policy_id"],
                "title": policy["title"],
                "source_row": policy["source_row"],
            }
            for policy in policies
        ],
        "prompt_tokens": usage.get("prompt_tokens"),
        "completion_tokens": usage.get("completion_tokens"),
        "total_tokens": usage.get("total_tokens"),
    }


def dumps_query_plan(query_plan: dict[str, Any]) -> str:
    return json.dumps(query_plan, ensure_ascii=False)
